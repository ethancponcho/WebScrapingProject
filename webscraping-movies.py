
from urllib.request import urlopen, Request
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


wb = xl.Workbook()

ws = wb.active

ws.title = 'Box Office Report'

ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Total Gross'
ws['E1'] = 'Theaters'
ws['F1'] = 'Average per theater'

#dimensions/font
first_row = ws.min_row
ws.cell(first_row,1).font= Font(size=16, bold=True)
ws.cell(first_row,2).font= Font(size=16, bold=True)
ws.cell(first_row,3).font= Font(size=16, bold=True)
ws.cell(first_row,4).font= Font(size=16, bold=True)
ws.cell(first_row,5).font= Font(size=16, bold=True)
ws.cell(first_row,6).font= Font(size=16, bold=True)

ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 25


#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2024/'
headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.3'}
req = Request(webpage, headers=headers)

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##print name, total gross, release date, average gross per theater, theaters, rank
##

movie_table = soup.find_all("td")




counter = 0
for x in range(5):
    rank = movie_table[counter].text
    name = movie_table[counter+1].text
    date = movie_table[counter+8].text
    total_gross = float(movie_table[counter+7].text.strip('$').replace(',',''))
    theaters = float(movie_table[counter+6].text.replace(',',''))
    avg_theater = total_gross / theaters
    
    counter +=11
    
    ws.append([rank, name, date, total_gross, theaters, avg_theater])


for cell in ws["D:D"]:
    cell.number_format = u'"$"#,##0.00'
for cell in ws["E:E"]:
    cell.number_format = '#,##0'
for cell in ws["F:F"]:
    cell.number_format = u'"$"#,##0.00'
        
wb.save('BoxOfficeReport.xlsx')